import os, shutil
from copy import copy
from datetime import datetime
from openpyxl import load_workbook
from services.normalizer_service import normalize_barcode, normalize_product, normalize_text, normalize_lote, normalize_date, make_key
from services.supabase_storage_service import upload_excel

BASE_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
MODELO_PATH = os.path.join(BASE_DIR, "modelos", "produtos_2026-05-29.xlsx")
DATA_DIR = os.path.join(BASE_DIR, "data")
BACKUP_DIR = os.path.join(BASE_DIR, "backups")
EXCEL_ATUAL_PATH = os.path.join(DATA_DIR, "produtos_atualizado.xlsx")

SHEET_NAME = "Relatório"
HEADER_ROW = 4
FIRST_DATA_ROW = 5

EXPECTED_HEADERS = {
    "produto": ["PRODUTO"],
    "categoria": ["CATEGORIA"],
    "local": ["LOCAL"],
    "codigo_barras": ["CÓDIGO DE BARRAS", "CODIGO DE BARRAS", "BARRAS", "EAN", "GTIN"],
    "lote": ["LOTE"],
    "validade": ["VENCIMENTO", "VALIDADE"],
    "aberto_em": ["ABERTO EM"],
    "vence_apos_aberto": ["VENCE APÓS ABERTO", "VENCE APOS ABERTO"],
    "estoque": ["ESTOQUE"],
    "estoque_padrao": ["ESTOQUE PADRÃO", "ESTOQUE PADRAO"],
    "limite_alerta": ["LIMITE ALERTA"],
    "status": ["STATUS"],
}

def ensure_workbook_exists():
    os.makedirs(DATA_DIR, exist_ok=True)
    os.makedirs(BACKUP_DIR, exist_ok=True)
    if not os.path.exists(EXCEL_ATUAL_PATH):
        if os.path.exists(MODELO_PATH):
            shutil.copy(MODELO_PATH, EXCEL_ATUAL_PATH)
        else:
            create_default_workbook(EXCEL_ATUAL_PATH)

def create_default_workbook(path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws["A1"] = "Controle de Estoque"
    ws["A2"] = "Relatório geral de produtos"
    headers = ["Produto", "Categoria", "Local", "Código de Barras", "Lote", "Vencimento", "Aberto em", "Vence após aberto", "Estoque", "Estoque padrão", "Limite alerta", "Status"]
    for col, value in enumerate(headers, start=1):
        ws.cell(row=HEADER_ROW, column=col).value = value
    wb.save(path)

def create_backup():
    if not os.path.exists(EXCEL_ATUAL_PATH):
        return None
    path = os.path.join(BACKUP_DIR, f"produtos_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    shutil.copy(EXCEL_ATUAL_PATH, path)
    return path

def detect_headers(ws):
    mapping = {}
    for cell in ws[HEADER_ROW]:
        value = str(cell.value or "").upper().strip()
        for field, aliases in EXPECTED_HEADERS.items():
            if any(alias in value for alias in aliases):
                mapping[field] = cell.column
    required = ["produto", "categoria", "local", "codigo_barras", "lote", "validade", "estoque", "estoque_padrao", "limite_alerta"]
    missing = [f for f in required if f not in mapping]
    if missing:
        raise ValueError(f"Cabeçalhos não encontrados no Excel: {', '.join(missing)}")
    return mapping

def build_existing_index(ws, hm):
    existing = {}
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        item = {
            "produto": ws.cell(row=row, column=hm["produto"]).value,
            "lote": ws.cell(row=row, column=hm["lote"]).value,
        }
        key = make_key(item)
        if all(key):
            existing[key] = row
    return existing

def clean_item(item):
    return {
        "produto": normalize_product(item.get("produto", "")),
        "categoria": normalize_text(item.get("categoria", "")),
        "local": normalize_text(item.get("local", "")),
        "codigo_barras": normalize_barcode(item.get("codigo_barras", "")),
        "lote": normalize_lote(item.get("lote", "")),
        "validade": normalize_date(item.get("validade", "")),
        "quantidade": int(item.get("quantidade", 0)),
        "estoque_padrao": int(item.get("estoque_padrao", 0)),
        "limite_alerta": int(item.get("limite_alerta", 0)),
    }

def save_items_to_excel(itens):
    ensure_workbook_exists()
    backup = create_backup()
    wb = load_workbook(EXCEL_ATUAL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    hm = detect_headers(ws)
    existing = build_existing_index(ws, hm)
    adicionados = 0
    atualizados = 0
    for item in itens:
        clean = clean_item(item)
        key = make_key(clean)
        if key in existing:
            row = existing[key]
            try:
                atual = int(ws.cell(row=row, column=hm["estoque"]).value or 0)
            except Exception:
                atual = 0
            ws.cell(row=row, column=hm["estoque"]).value = atual + clean["quantidade"]
            write_common_fields(ws, row, hm, clean)
            atualizados += 1
        else:
            row = find_next_empty_row(ws, hm)
            copy_style_from_previous_row(ws, row)
            write_item(ws, row, hm, clean)
            existing[key] = row
            adicionados += 1
    wb.save(EXCEL_ATUAL_PATH)

    supabase_ok, supabase_msg = upload_excel(EXCEL_ATUAL_PATH)
    print(f"[Supabase] {supabase_msg}")

    return {
        "adicionados": adicionados,
        "atualizados": atualizados,
        "backup": backup,
        "supabase_ok": supabase_ok,
        "supabase_msg": supabase_msg,
    }

def find_product_by_barcode(codigo):
    ensure_workbook_exists()
    wb = load_workbook(EXCEL_ATUAL_PATH, data_only=True)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    hm = detect_headers(ws)
    codigo_limpo = normalize_barcode(codigo)
    encontrado = None
    for row in range(FIRST_DATA_ROW, ws.max_row + 1):
        cod = normalize_barcode(ws.cell(row=row, column=hm["codigo_barras"]).value)
        if cod and cod == codigo_limpo:
            encontrado = {
                "codigo_barras": codigo_limpo,
                "produto": str(ws.cell(row=row, column=hm["produto"]).value or ""),
                "categoria": str(ws.cell(row=row, column=hm["categoria"]).value or ""),
                "local": str(ws.cell(row=row, column=hm["local"]).value or ""),
                "lote": str(ws.cell(row=row, column=hm["lote"]).value or ""),
                "validade": str(ws.cell(row=row, column=hm["validade"]).value or ""),
                "estoque_padrao": str(ws.cell(row=row, column=hm["estoque_padrao"]).value or ""),
                "limite_alerta": str(ws.cell(row=row, column=hm["limite_alerta"]).value or ""),
            }
    return encontrado

def find_next_empty_row(ws, hm):
    for row in range(FIRST_DATA_ROW, ws.max_row + 2):
        if ws.cell(row=row, column=hm["produto"]).value in [None, ""]:
            return row
    return ws.max_row + 1

def copy_style_from_previous_row(ws, row):
    source_row = row - 1 if row > FIRST_DATA_ROW else FIRST_DATA_ROW
    for col in range(1, ws.max_column + 1):
        source = ws.cell(row=source_row, column=col)
        target = ws.cell(row=row, column=col)
        if source.has_style:
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.font = copy(source.font)
            target.fill = copy(source.fill)
            target.border = copy(source.border)
            target.alignment = copy(source.alignment)

def set_if_exists(ws, row, hm, field, value):
    if field in hm:
        ws.cell(row=row, column=hm[field]).value = value

def write_common_fields(ws, row, hm, item):
    set_if_exists(ws, row, hm, "categoria", item["categoria"])
    set_if_exists(ws, row, hm, "local", item["local"])
    if item["codigo_barras"]:
        cell = ws.cell(row=row, column=hm["codigo_barras"])
        cell.value = str(item["codigo_barras"])
        cell.number_format = "@"
    set_if_exists(ws, row, hm, "validade", item["validade"])
    set_if_exists(ws, row, hm, "estoque_padrao", item["estoque_padrao"])
    set_if_exists(ws, row, hm, "limite_alerta", item["limite_alerta"])

def write_item(ws, row, hm, item):
    set_if_exists(ws, row, hm, "produto", item["produto"])
    write_common_fields(ws, row, hm, item)
    if "codigo_barras" in hm and not item["codigo_barras"]:
        ws.cell(row=row, column=hm["codigo_barras"]).value = "-"
    set_if_exists(ws, row, hm, "lote", item["lote"])
    set_if_exists(ws, row, hm, "estoque", int(item["quantidade"]))
    set_if_exists(ws, row, hm, "aberto_em", "")
    set_if_exists(ws, row, hm, "vence_apos_aberto", "")
    set_if_exists(ws, row, hm, "status", "")
