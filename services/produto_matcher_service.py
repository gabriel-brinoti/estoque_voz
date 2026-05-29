from difflib import SequenceMatcher
from openpyxl import load_workbook
from services.excel_service import ensure_workbook_exists, EXCEL_ATUAL_PATH, SHEET_NAME, FIRST_DATA_ROW, detect_headers
from services.normalizer_service import normalize_product, normalize_product_for_match

MIN_SCORE = 0.72

def load_product_library():
    ensure_workbook_exists()
    wb=load_workbook(EXCEL_ATUAL_PATH,data_only=True)
    ws=wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    hm=detect_headers(ws)
    produtos=[]; vistos=set()
    for row in range(FIRST_DATA_ROW, ws.max_row+1):
        produto=normalize_product(ws.cell(row=row,column=hm["produto"]).value)
        if produto and produto not in vistos:
            produtos.append(produto); vistos.add(produto)
    return produtos

def similarity(a,b):
    a=normalize_product_for_match(a); b=normalize_product_for_match(b)
    if not a or not b: return 0
    base=SequenceMatcher(None,a,b).ratio()
    aw=set(a.split()); bw=set(b.split())
    word=len(aw & bw)/len(aw | bw) if aw and bw else 0
    return base*0.75 + word*0.25

def suggest_product_name(spoken_name):
    spoken=normalize_product(spoken_name)
    library=load_product_library()
    best=""; score=0
    for product in library:
        s=similarity(spoken,product)
        if s>score: best=product; score=s
    found=score>=MIN_SCORE
    return {"found":found,"produto_original":spoken,"suggestion":best if found else "","score":round(score,2),"message":"Sugestão encontrada." if found else "Nenhuma sugestão confiável."}
